"""
Accuracy report generator: compares FRAS system output against a manual ground truth CSV
and produces an Excel report with accuracy, false positive, and false negative stats.

Usage:
    python src/report.py --ground-truth data/ground_truth.csv --session morning_01

Ground truth CSV format: student_id, date, session, actual_status (Present/Absent)
"""

import argparse
import os
import sqlite3
from datetime import datetime

import pandas as pd
import yaml
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def load_config(config_path: str = "config.yaml") -> dict:
    with open(config_path) as f:
        return yaml.safe_load(f)


def load_system_records(db_path: str, session: str | None = None) -> pd.DataFrame:
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Attendance DB not found: {db_path}")
    conn = sqlite3.connect(db_path)
    query = "SELECT student_id, date, session, status as system_status FROM attendance"
    if session:
        query += f" WHERE session = '{session}'"
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df


def generate_report(ground_truth_path: str, config: dict, session: str | None = None) -> str:
    db_path = config["paths"]["attendance_db"]

    ground_truth = pd.read_csv(ground_truth_path)
    ground_truth.columns = [c.strip().lower() for c in ground_truth.columns]
    ground_truth = ground_truth.rename(columns={"actual_status": "actual"})

    system = load_system_records(db_path, session)
    system = system.rename(columns={"status": "system_status"})

    merged = pd.merge(ground_truth, system, on=["student_id", "date", "session"], how="left")
    merged["system_status"] = merged["system_status"].fillna("Absent")

    # Per-student comparison
    merged["correct"] = merged["actual"] == merged["system_status"]
    merged["false_positive"] = (merged["actual"] == "Absent") & (merged["system_status"] == "Present")
    merged["false_negative"] = (merged["actual"] == "Present") & (merged["system_status"] == "Absent")

    total = len(merged)
    correct = merged["correct"].sum()
    fp = merged["false_positive"].sum()
    fn = merged["false_negative"].sum()

    accuracy = round(correct / total * 100, 2) if total else 0
    fp_rate = round(fp / total * 100, 2) if total else 0
    fn_rate = round(fn / total * 100, 2) if total else 0

    # Success criteria check (from Scope of Work)
    criteria = {
        "Identification Accuracy ≥ 85%": (accuracy >= 85, f"{accuracy}%"),
        "False Positives < 2%": (fp_rate < 2, f"{fp_rate}%"),
    }

    print(f"\n{'='*40}")
    print(f"  FRAS PILOT ACCURACY REPORT")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*40}")
    print(f"  Total records : {total}")
    print(f"  Accuracy      : {accuracy}%")
    print(f"  False Positives: {fp_rate}%")
    print(f"  False Negatives: {fn_rate}%")
    print(f"\n  Pilot Success Criteria:")
    for criterion, (passed, value) in criteria.items():
        status = "PASS" if passed else "FAIL"
        print(f"  [{status}] {criterion} → {value}")
    print(f"{'='*40}\n")

    # Write Excel report
    import openpyxl
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    summary_data = [
        ["FRAS Pilot Accuracy Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Session", session or "All"],
        ["", ""],
        ["Metric", "Value"],
        ["Total Records", total],
        ["Accuracy", f"{accuracy}%"],
        ["False Positive Rate", f"{fp_rate}%"],
        ["False Negative Rate", f"{fn_rate}%"],
        ["", ""],
        ["Pilot Success Criteria", ""],
    ]
    for criterion, (passed, value) in criteria.items():
        summary_data.append([criterion, f"{'PASS' if passed else 'FAIL'} ({value})"])

    for row in summary_data:
        ws1.append(row)

    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A5"].font = Font(bold=True)
    ws1["A11"].font = Font(bold=True)

    # Sheet 2: Per-student detail
    ws2 = wb.create_sheet("Detail")
    detail_cols = ["student_id", "date", "session", "actual", "system_status", "correct", "false_positive", "false_negative"]
    for r in dataframe_to_rows(merged[detail_cols], index=False, header=True):
        ws2.append(r)

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"reports/fras_pilot_report_{timestamp}.xlsx"
    wb.save(output_path)
    print(f"[report] Saved → {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ground-truth", required=True, help="Path to manual ground truth CSV")
    parser.add_argument("--session", default=None, help="Filter to a specific session label")
    parser.add_argument("--config", default="config.yaml")
    args = parser.parse_args()

    config = load_config(args.config)
    generate_report(args.ground_truth, config, args.session)
